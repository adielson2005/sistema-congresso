import os
import re
import sqlite3
import bcrypt
from datetime import datetime, timedelta
from time import time
from secrets import compare_digest, token_hex, token_urlsafe
from io import BytesIO
from urllib.parse import urlparse
from flask import Flask, render_template, request, redirect, session, flash, jsonify, make_response
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from xhtml2pdf import pisa
from werkzeug.middleware.proxy_fix import ProxyFix

# =========================
# CONFIG
# =========================
DB_PATH = os.environ.get("DB_PATH", "database/banco.db")
db_dir = os.path.dirname(DB_PATH)
if db_dir:
    os.makedirs(db_dir, exist_ok=True)

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

APP_ENV = os.environ.get("APP_ENV", "development").strip().lower()
IS_PRODUCTION = APP_ENV == "production"

secret_key = os.environ.get("SECRET_KEY")
if IS_PRODUCTION and not secret_key:
    raise RuntimeError("SECRET_KEY é obrigatório com APP_ENV=production")
app.secret_key = secret_key or token_hex(32)

session_cookie_secure_env = os.environ.get("SESSION_COOKIE_SECURE", "")
session_cookie_secure = (
    session_cookie_secure_env.lower() in {"1", "true", "yes"}
    if session_cookie_secure_env
    else IS_PRODUCTION
)

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = session_cookie_secure
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=int(os.environ.get("SESSION_LIFETIME_HOURS", "12")))
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_CONTENT_LENGTH", str(2 * 1024 * 1024)))
app.config["PREFERRED_URL_SCHEME"] = "https" if IS_PRODUCTION else "http"

UNSAFE_METHODS = {"POST", "PUT", "PATCH", "DELETE"}
DEMO_ALLOWED_ENDPOINTS = {"login", "demo", "logout"}
PARTICIPANTES_POR_PAGINA = 10
DEMO_PASSWORD = os.environ.get("DEMO_PASSWORD")
SINGLE_LOGIN_USER = os.environ.get("SINGLE_LOGIN_USER")
SINGLE_LOGIN_PASSWORD = os.environ.get("SINGLE_LOGIN_PASSWORD")
SINGLE_LOGIN_NAME = os.environ.get("SINGLE_LOGIN_NAME", "Login compartilhado")
SINGLE_LOGIN_ROLE = os.environ.get("SINGLE_LOGIN_ROLE", "admin")
LOGIN_MAX_ATTEMPTS = int(os.environ.get("LOGIN_MAX_ATTEMPTS", "10"))
LOGIN_ATTEMPT_WINDOW_SECONDS = int(os.environ.get("LOGIN_ATTEMPT_WINDOW_SECONDS", "900"))
LOGIN_LOCKOUT_SECONDS = int(os.environ.get("LOGIN_LOCKOUT_SECONDS", "900"))
FAILED_LOGIN_ATTEMPTS = {}


# =========================
# BANCO
# =========================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_client_ip():
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr or "unknown"


def is_safe_target(target):
    if not target:
        return False

    host_url = urlparse(request.host_url)
    redirect_url = urlparse(target)
    if redirect_url.scheme and redirect_url.scheme not in {"http", "https"}:
        return False
    return (not redirect_url.netloc) or (redirect_url.netloc == host_url.netloc)


def safe_redirect_back(default_path):
    target = request.referrer
    if target and is_safe_target(target):
        return redirect(target)
    return redirect(default_path)


def prune_failed_attempts(now_ts):
    keys_to_delete = []
    for ip_address, info in FAILED_LOGIN_ATTEMPTS.items():
        attempts = [ts for ts in info.get("attempts", []) if now_ts - ts <= LOGIN_ATTEMPT_WINDOW_SECONDS]
        info["attempts"] = attempts
        if info.get("locked_until", 0) < now_ts and not attempts:
            keys_to_delete.append(ip_address)

    for key in keys_to_delete:
        FAILED_LOGIN_ATTEMPTS.pop(key, None)


def is_ip_locked(ip_address, now_ts):
    info = FAILED_LOGIN_ATTEMPTS.get(ip_address)
    if not info:
        return False

    locked_until = info.get("locked_until", 0)
    return locked_until > now_ts


def register_failed_login(ip_address, now_ts):
    info = FAILED_LOGIN_ATTEMPTS.setdefault(ip_address, {"attempts": [], "locked_until": 0})
    attempts = [ts for ts in info["attempts"] if now_ts - ts <= LOGIN_ATTEMPT_WINDOW_SECONDS]
    attempts.append(now_ts)
    info["attempts"] = attempts

    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        info["locked_until"] = now_ts + LOGIN_LOCKOUT_SECONDS
        info["attempts"] = []


def clear_failed_login(ip_address):
    FAILED_LOGIN_ATTEMPTS.pop(ip_address, None)


def get_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def is_valid_csrf_request():
    session_token = session.get("_csrf_token")
    request_token = request.form.get("_csrf_token") or request.headers.get("X-CSRF-Token")
    return bool(session_token and request_token and compare_digest(session_token, request_token))


def parse_currency_input(value):
    normalized = (value or "").strip()
    if not normalized:
        raise ValueError("Informe um valor válido.")

    normalized = normalized.replace("R$", "").replace(" ", "")

    if "," in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")

    return float(normalized)


def only_digits(value):
    return re.sub(r"\D", "", value or "")


def format_cpf(value):
    digits = only_digits(value)
    if len(digits) != 11:
        return value
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def format_phone(value):
    digits = only_digits(value)
    if not digits:
        return ""
    if len(digits) == 10:
        return f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
    return value


def is_valid_cpf(value):
    digits = only_digits(value)

    if len(digits) != 11 or digits == digits[0] * 11:
        return False

    total = sum(int(digits[index]) * (10 - index) for index in range(9))
    first_digit = (total * 10) % 11
    if first_digit == 10:
        first_digit = 0

    total = sum(int(digits[index]) * (11 - index) for index in range(10))
    second_digit = (total * 10) % 11
    if second_digit == 10:
        second_digit = 0

    return digits[-2:] == f"{first_digit}{second_digit}"


def is_valid_email(value):
    email = (value or "").strip()
    if not email:
        return True
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email))


def email_already_exists(conn, email, exclude_id=None):
    email_value = (email or "").strip()
    if not email_value:
        return False

    if exclude_id is None:
        row = conn.execute(
            """
            SELECT 1
            FROM participantes
            WHERE LOWER(email) = LOWER(?)
            LIMIT 1
            """,
            (email_value,),
        ).fetchone()
        return bool(row)

    row = conn.execute(
        """
        SELECT 1
        FROM participantes
        WHERE LOWER(email) = LOWER(?)
          AND id != ?
        LIMIT 1
        """,
        (email_value, exclude_id),
    ).fetchone()
    return bool(row)


def validate_iso_date(value, field_name):
    try:
        datetime.strptime((value or "").strip(), "%Y-%m-%d")
    except ValueError as error:
        raise ValueError(f"{field_name} inválida.") from error


def normalize_participante_fields(cpf, numero, email, data_nascimento):
    if not is_valid_cpf(cpf):
        raise ValueError("CPF inválido.")

    if not is_valid_email(email):
        raise ValueError("E-mail inválido.")

    validate_iso_date(data_nascimento, "Data de nascimento")

    phone_digits = only_digits(numero)
    if phone_digits and len(phone_digits) not in {10, 11}:
        raise ValueError("Telefone inválido. Use 10 ou 11 dígitos.")

    return format_cpf(cpf), format_phone(numero)


def normalize_existing_participantes(cursor):
    participantes = cursor.execute("""
        SELECT id, cpf, numero
        FROM participantes
    """).fetchall()

    for participante_id, cpf, numero in participantes:
        cpf_digits = only_digits(cpf)
        phone_digits = only_digits(numero)

        if not cpf_digits or not is_valid_cpf(cpf_digits):
            continue

        if phone_digits and len(phone_digits) not in {10, 11}:
            continue

        cpf_formatado = format_cpf(cpf_digits)
        numero_formatado = format_phone(phone_digits)

        if cpf != cpf_formatado or (numero or "") != numero_formatado:
            cursor.execute("""
                UPDATE participantes
                SET cpf = ?, numero = ?
                WHERE id = ?
            """, (cpf_formatado, numero_formatado, participante_id))


@app.context_processor
def inject_template_helpers():
    return {"csrf_token": get_csrf_token}


@app.before_request
def protect_unsafe_requests():
    if request.method not in UNSAFE_METHODS:
        return None

    if not is_valid_csrf_request():
        flash("Sessão inválida ou expirada. Atualize a página e tente novamente.", "danger")
        return safe_redirect_back("/login")

    if session.get("modo_demo") and request.endpoint not in DEMO_ALLOWED_ENDPOINTS:
        flash("Modo demonstração: alterações estão bloqueadas.", "warning")
        return safe_redirect_back("/dashboard")

    return None


@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    if IS_PRODUCTION and request.is_secure:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


def authenticate_single_login(usuario, senha):
    if not SINGLE_LOGIN_USER or not SINGLE_LOGIN_PASSWORD:
        return None

    if not compare_digest(usuario, SINGLE_LOGIN_USER):
        return False

    if not compare_digest(senha, SINGLE_LOGIN_PASSWORD):
        return False

    return {
        "id": 1,
        "nome": SINGLE_LOGIN_NAME,
        "cargo": SINGLE_LOGIN_ROLE,
    }


def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            usuario TEXT UNIQUE,
            senha_hash BLOB,
            cargo TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS macros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS congregacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            macro_id INTEGER,
            UNIQUE(nome, macro_id),
            FOREIGN KEY (macro_id) REFERENCES macros(id)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS participantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_completo TEXT,
            data_nascimento TEXT,
            cpf TEXT UNIQUE,
            email TEXT,
            numero TEXT,
            nome_mae TEXT,
            congregacao TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS arrecadacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            participante_id INTEGER,
            valor REAL,
            data_lancamento TEXT,
            observacao TEXT,
            FOREIGN KEY (participante_id) REFERENCES participantes(id)
        )
    """)

    default_admin_user = os.environ.get("DEFAULT_ADMIN_USER")
    default_admin_password = os.environ.get("DEFAULT_ADMIN_PASSWORD")
    default_admin_name = os.environ.get("DEFAULT_ADMIN_NAME", "Administrador")
    default_admin_role = os.environ.get("DEFAULT_ADMIN_ROLE", "admin")

    if default_admin_user and default_admin_password:
        user = cursor.execute("""
            SELECT * FROM usuarios WHERE usuario = ?
        """, (default_admin_user,)).fetchone()

        if not user:
            senha_hash = bcrypt.hashpw(default_admin_password.encode("utf-8"), bcrypt.gensalt())
            cursor.execute("""
                INSERT INTO usuarios (nome, usuario, senha_hash, cargo)
                VALUES (?, ?, ?, ?)
            """, (default_admin_name, default_admin_user, senha_hash, default_admin_role))

    # macros padrão
    macros = [
        "MACRO 1",
        "MACRO 2",
        "MACRO 3",
        "MACRO 4",
        "MACRO 5",
        "MACRO 6",
        "MACRO MISSª"
    ]

    for macro in macros:
        cursor.execute("INSERT OR IGNORE INTO macros (nome) VALUES (?)", (macro,))

    cursor.execute("SELECT id, nome FROM macros")
    macro_map = {row[1]: row[0] for row in cursor.fetchall()}

    dados_congregacoes = {
        "MACRO 1": [
            "Betesda", "El-Raah", "E. do Altíssimo",
            "Morada do Altíssimo", "Hebrom", "Shalom"
        ],
        "MACRO 2": [
            "Alto Refúgio", "Jardim de Deus", "Jardim Celeste",
            "Manancial de Vida", "Monte Carmelo", "Orvalho de Hermon",
            "Rosa de Sarom", "Cafarnaum", "Ebenezer", "Monte Ararate"
        ],
        "MACRO 3": [
            "Betânia", "Deus Forte", "Monte Hermon",
            "Monte Santo", "Monte Sinai", "Rio Jordão"
        ],
        "MACRO 4": [
            "Adonai", "Fonte de Luz", "Gileade",
            "Rocha Eterna", "Vitória da Fé", "Monte Horebe",
            "Nova Jerusalém"
        ],
        "MACRO 5": [
            "Monte das Oliveiras", "Monte Sião", "Nova Canaã",
            "Porta do Céu", "Nova Aliança", "Filadélfia",
            "Fonte de Águas Vivas"
        ],
        "MACRO 6": [
            "Betel", "Monte Tabor", "Nova Sião", "Pioneira",
            "Monte Moriá", "Getsêmani", "Joia de Cristo",
            "Maranata", "Nova Vida", "Porta Formosa"
        ],
        "MACRO MISSª": [
            "Luz do Mundo", "Vale de Benção", "Lírio dos Vales"
        ]
    }

    for macro_nome, congregacoes in dados_congregacoes.items():
        macro_id = macro_map.get(macro_nome)
        if macro_id:
            for nome in congregacoes:
                cursor.execute("""
                    INSERT OR IGNORE INTO congregacoes (nome, macro_id)
                    VALUES (?, ?)
                """, (nome, macro_id))

    normalize_existing_participantes(cursor)

    conn.commit()
    conn.close()


init_db()


# =========================
# ROTAS BÁSICAS
# =========================
@app.route("/")
def home():
    return redirect("/login")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        now_ts = int(time())
        prune_failed_attempts(now_ts)
        client_ip = get_client_ip()

        if is_ip_locked(client_ip, now_ts):
            flash("Muitas tentativas. Aguarde alguns minutos e tente novamente.", "danger")
            return redirect("/login")

        usuario = request.form["usuario"]
        senha = request.form["senha"]

        single_login_user = authenticate_single_login(usuario, senha)
        if single_login_user:
            clear_failed_login(client_ip)
            session.pop("modo_demo", None)
            session["user_id"] = single_login_user["id"]
            session["nome"] = single_login_user["nome"]
            session["cargo"] = single_login_user["cargo"]
            session.permanent = True
            return redirect("/dashboard")
        if single_login_user is False:
            register_failed_login(client_ip, now_ts)
            flash("Usuário ou senha inválidos.", "danger")
            return redirect("/login")

        conn = get_db()
        user = conn.execute("""
            SELECT * FROM usuarios
            WHERE usuario = ?
        """, (usuario,)).fetchone()
        conn.close()

        if user:
            senha_hash = user["senha_hash"]
            if bcrypt.checkpw(senha.encode("utf-8"), senha_hash):
                clear_failed_login(client_ip)
                session.pop("modo_demo", None)
                session["user_id"] = user["id"]
                session["nome"] = user["nome"]
                session["cargo"] = user["cargo"]
                session.permanent = True
                return redirect("/dashboard")
            register_failed_login(client_ip, now_ts)
            flash("Usuário ou senha inválidos.", "danger")
            return redirect("/login")

        register_failed_login(client_ip, now_ts)
        flash("Usuário ou senha inválidos.", "danger")
        return redirect("/login")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# =========================
# DASHBOARD
# =========================
@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    total_participantes = conn.execute("""
        SELECT COUNT(*) AS total FROM participantes
    """).fetchone()["total"]

    total_arrecadado = conn.execute("""
        SELECT COALESCE(SUM(valor), 0) AS total FROM arrecadacoes
    """).fetchone()["total"]

    total_lancamentos = conn.execute("""
        SELECT COUNT(*) AS total FROM arrecadacoes
    """).fetchone()["total"]

    ultimas_arrecadacoes = conn.execute("""
        SELECT arrecadacoes.*, participantes.nome_completo
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        ORDER BY arrecadacoes.id DESC
        LIMIT 5
    """).fetchall()

    arrecadacao_por_congregacao = conn.execute("""
        SELECT participantes.congregacao, COALESCE(SUM(arrecadacoes.valor), 0) AS total
        FROM participantes
        LEFT JOIN arrecadacoes ON arrecadacoes.participante_id = participantes.id
        GROUP BY participantes.congregacao
        ORDER BY total DESC
        LIMIT 10
    """).fetchall()

    arrecadacao_por_macro = conn.execute("""
        SELECT macros.nome AS macro, COALESCE(SUM(arrecadacoes.valor), 0) AS total
        FROM macros
        LEFT JOIN congregacoes ON congregacoes.macro_id = macros.id
        LEFT JOIN participantes ON participantes.congregacao = congregacoes.nome
        LEFT JOIN arrecadacoes ON arrecadacoes.participante_id = participantes.id
        GROUP BY macros.id, macros.nome
        ORDER BY macros.nome
    """).fetchall()

    evolucao_arrecadacao = conn.execute("""
        SELECT data_lancamento, COALESCE(SUM(valor), 0) AS total
        FROM arrecadacoes
        GROUP BY data_lancamento
        ORDER BY data_lancamento ASC
    """).fetchall()

    conn.close()

    return render_template(
        "dashboard.html",
        total_participantes=total_participantes,
        total_arrecadado=total_arrecadado,
        total_lancamentos=total_lancamentos,
        ultimas_arrecadacoes=ultimas_arrecadacoes,
        arrecadacao_por_congregacao=arrecadacao_por_congregacao,
        arrecadacao_por_macro=arrecadacao_por_macro,
        evolucao_arrecadacao=evolucao_arrecadacao
    )


@app.route("/api/dashboard/evolucao")
def api_dashboard_evolucao():
    if "user_id" not in session:
        return jsonify([])

    conn = get_db()
    evolucao = conn.execute("""
        SELECT data_lancamento, COALESCE(SUM(valor), 0) AS total
        FROM arrecadacoes
        GROUP BY data_lancamento
        ORDER BY data_lancamento ASC
    """).fetchall()
    conn.close()

    return jsonify([
        {"data": item["data_lancamento"], "total": float(item["total"])}
        for item in evolucao
    ])


# =========================
# PARTICIPANTES
# =========================
@app.route("/participantes")
def participantes():
    if "user_id" not in session:
        return redirect("/login")

    busca = request.args.get("busca", "").strip()
    macro = request.args.get("macro", "").strip()
    congregacao = request.args.get("congregacao", "").strip()
    page = max(request.args.get("page", 1, type=int), 1)

    conn = get_db()

    query_base = "FROM participantes WHERE 1=1"
    params = []

    if busca:
        query_base += """
            AND (
                nome_completo LIKE ?
                OR cpf LIKE ?
                OR email LIKE ?
                OR numero LIKE ?
                OR nome_mae LIKE ?
                OR congregacao LIKE ?
            )
        """
        termo = f"%{busca}%"
        params.extend([termo, termo, termo, termo, termo, termo])

    if macro:
        query_base += " AND congregacao IN (SELECT nome FROM congregacoes WHERE macro_id = ?)"
        params.append(macro)

    if congregacao:
        query_base += " AND congregacao = ?"
        params.append(congregacao)

    total_participantes = conn.execute(
        f"SELECT COUNT(*) AS total {query_base}",
        params,
    ).fetchone()["total"]

    total_paginas = max((total_participantes + PARTICIPANTES_POR_PAGINA - 1) // PARTICIPANTES_POR_PAGINA, 1)
    if page > total_paginas:
        page = total_paginas

    offset = (page - 1) * PARTICIPANTES_POR_PAGINA
    query = f"SELECT * {query_base} ORDER BY id DESC LIMIT ? OFFSET ?"
    query_params = params + [PARTICIPANTES_POR_PAGINA, offset]

    lista_participantes = conn.execute(query, query_params).fetchall()

    macros = conn.execute("""
        SELECT * FROM macros
        ORDER BY nome
    """).fetchall()

    congregacoes = conn.execute("""
        SELECT DISTINCT nome
        FROM congregacoes
        ORDER BY nome
    """).fetchall()

    conn.close()

    return render_template(
        "participantes.html",
        participantes=lista_participantes,
        busca=busca,
        macros=macros,
        congregacoes=congregacoes,
        macro_selecionada=macro,
        congregacao_selecionada=congregacao,
        pagina_atual=page,
        total_paginas=total_paginas,
        total_participantes=total_participantes
    )


@app.route("/participantes/cadastrar", methods=["GET", "POST"])
def cadastrar_participante():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    if request.method == "POST":
        nome_completo = request.form["nome_completo"]
        data_nascimento = request.form["data_nascimento"]
        cpf = request.form["cpf"]
        email = request.form["email"].strip()
        numero = request.form["numero"]
        nome_mae = request.form["nome_mae"]
        congregacao = request.form["congregacao"]

        try:
            cpf, numero = normalize_participante_fields(cpf, numero, email, data_nascimento)
        except ValueError as error:
            conn.close()
            flash(str(error), "danger")
            return redirect("/participantes/cadastrar")

        if email_already_exists(conn, email):
            conn.close()
            flash("E-mail já cadastrado para outro participante.", "danger")
            return redirect("/participantes/cadastrar")

        try:
            conn.execute("""
                INSERT INTO participantes (
                    nome_completo, data_nascimento, cpf, email, numero, nome_mae, congregacao
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                nome_completo,
                data_nascimento,
                cpf,
                email,
                numero,
                nome_mae,
                congregacao
            ))
            conn.commit()
            flash("Participante cadastrado com sucesso!", "success")
        except sqlite3.IntegrityError:
            conn.close()
            flash("CPF já cadastrado. Verifique os dados.", "danger")
            return redirect("/participantes/cadastrar")

        conn.close()
        return redirect("/participantes")

    macros = conn.execute("""
        SELECT * FROM macros
        ORDER BY nome
    """).fetchall()

    conn.close()
    return render_template("cadastrar_participante.html", macros=macros)


@app.route("/participantes/<int:id>")
def detalhe_participante(id):
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    participante = conn.execute("""
        SELECT * FROM participantes
        WHERE id = ?
    """, (id,)).fetchone()

    arrecadacoes = conn.execute("""
        SELECT * FROM arrecadacoes
        WHERE participante_id = ?
        ORDER BY id DESC
    """, (id,)).fetchall()

    total = conn.execute("""
        SELECT COALESCE(SUM(valor), 0) AS total_arrecadado
        FROM arrecadacoes
        WHERE participante_id = ?
    """, (id,)).fetchone()

    conn.close()

    if not participante:
        return "Participante não encontrado"

    return render_template(
        "detalhe_participante.html",
        participante=participante,
        arrecadacoes=arrecadacoes,
        total=total["total_arrecadado"]
    )


@app.route("/participantes/<int:id>/editar", methods=["GET", "POST"])
def editar_participante(id):
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()
    participante = conn.execute("""
        SELECT * FROM participantes
        WHERE id = ?
    """, (id,)).fetchone()

    if not participante:
        conn.close()
        flash("Participante não encontrado.", "danger")
        return redirect("/participantes")

    if request.method == "POST":
        nome_completo = request.form["nome_completo"]
        data_nascimento = request.form["data_nascimento"]
        cpf = request.form["cpf"]
        email = request.form["email"].strip()
        numero = request.form["numero"]
        nome_mae = request.form["nome_mae"]
        congregacao = request.form["congregacao"]

        try:
            cpf, numero = normalize_participante_fields(cpf, numero, email, data_nascimento)
        except ValueError as error:
            conn.close()
            flash(str(error), "danger")
            return redirect(f"/participantes/{id}/editar")

        if email_already_exists(conn, email, exclude_id=id):
            conn.close()
            flash("E-mail já cadastrado para outro participante.", "danger")
            return redirect(f"/participantes/{id}/editar")

        try:
            conn.execute("""
                UPDATE participantes
                SET nome_completo = ?, data_nascimento = ?, cpf = ?, email = ?, numero = ?, nome_mae = ?, congregacao = ?
                WHERE id = ?
            """, (
                nome_completo,
                data_nascimento,
                cpf,
                email,
                numero,
                nome_mae,
                congregacao,
                id
            ))
            conn.commit()
            flash("Participante atualizado com sucesso!", "success")
        except sqlite3.IntegrityError:
            conn.close()
            flash("CPF já cadastrado em outro participante.", "danger")
            return redirect(f"/participantes/{id}/editar")

        conn.close()
        return redirect(f"/participantes/{id}")

    macros = conn.execute("""
        SELECT * FROM macros
        ORDER BY nome
    """).fetchall()

    macro_atual = conn.execute("""
        SELECT macro_id
        FROM congregacoes
        WHERE nome = ?
        LIMIT 1
    """, (participante["congregacao"],)).fetchone()

    conn.close()

    return render_template(
        "editar_participante.html",
        participante=participante,
        macros=macros,
        macro_atual_id=macro_atual["macro_id"] if macro_atual else ""
    )


@app.route("/participantes/<int:id>/excluir", methods=["POST"])
def excluir_participante(id):
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()
    conn.execute("DELETE FROM arrecadacoes WHERE participante_id = ?", (id,))
    conn.execute("DELETE FROM participantes WHERE id = ?", (id,))
    conn.commit()
    conn.close()

    flash("Participante excluído com sucesso.", "success")
    return redirect("/participantes")


# =========================
# CONGREGAÇÕES
# =========================
@app.route("/api/congregacoes/<int:macro_id>")
def api_congregacoes(macro_id):
    if "user_id" not in session:
        return jsonify([])

    conn = get_db()
    congregacoes = conn.execute("""
        SELECT DISTINCT id, nome
        FROM congregacoes
        WHERE macro_id = ?
        ORDER BY nome
    """, (macro_id,)).fetchall()
    conn.close()

    return jsonify([
        {"id": c["id"], "nome": c["nome"]}
        for c in congregacoes
    ])


@app.route("/congregacoes")
def congregacoes():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    lista_congregacoes = conn.execute("""
        SELECT congregacoes.nome, macros.nome AS macro_nome,
               COUNT(participantes.id) AS total_participantes
        FROM congregacoes
        JOIN macros ON macros.id = congregacoes.macro_id
        LEFT JOIN participantes ON participantes.congregacao = congregacoes.nome
        GROUP BY congregacoes.id, congregacoes.nome, macros.nome
        ORDER BY macros.nome, congregacoes.nome
    """).fetchall()

    conn.close()

    return render_template("congregacoes.html", congregacoes=lista_congregacoes)


@app.route("/congregacoes/<nome_slug>")
def detalhe_congregacao(nome_slug):
    if "user_id" not in session:
        return redirect("/login")

    nome_congregacao = nome_slug.replace("-", " ")

    conn = get_db()

    participantes_da_congregacao = conn.execute("""
        SELECT * FROM participantes
        WHERE congregacao = ?
        ORDER BY nome_completo ASC
    """, (nome_congregacao,)).fetchall()

    congregacao_info = conn.execute("""
        SELECT congregacoes.nome, macros.nome AS macro_nome
        FROM congregacoes
        JOIN macros ON macros.id = congregacoes.macro_id
        WHERE congregacoes.nome = ?
    """, (nome_congregacao,)).fetchone()

    conn.close()

    if not congregacao_info:
        return "Congregação não encontrada."

    return render_template(
        "detalhe_congregacao.html",
        congregacao=congregacao_info["nome"],
        macro=congregacao_info["macro_nome"],
        participantes=participantes_da_congregacao
    )


# =========================
# AUDITORIA
# =========================
@app.route("/auditoria/inconsistencias")
def auditoria_inconsistencias():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    participantes = conn.execute("""
        SELECT id, nome_completo, cpf, email, numero, data_nascimento
        FROM participantes
        ORDER BY nome_completo ASC
    """).fetchall()

    emails_duplicados = {
        row["email_norm"]
        for row in conn.execute("""
            SELECT LOWER(TRIM(email)) AS email_norm, COUNT(*) AS total
            FROM participantes
            WHERE TRIM(COALESCE(email, '')) <> ''
            GROUP BY LOWER(TRIM(email))
            HAVING COUNT(*) > 1
        """).fetchall()
    }

    conn.close()

    inconsistencias = []
    for participante in participantes:
        problemas = []

        if not is_valid_cpf(participante["cpf"]):
            problemas.append("CPF inválido")

        phone_digits = only_digits(participante["numero"])
        if phone_digits and len(phone_digits) not in {10, 11}:
            problemas.append("Telefone inválido")

        if not is_valid_email(participante["email"]):
            problemas.append("E-mail inválido")

        email_normalizado = (participante["email"] or "").strip().lower()
        if email_normalizado and email_normalizado in emails_duplicados:
            problemas.append("E-mail duplicado")

        try:
            validate_iso_date(participante["data_nascimento"], "Data de nascimento")
        except ValueError:
            problemas.append("Data de nascimento inválida")

        if problemas:
            inconsistencias.append({
                "id": participante["id"],
                "nome": participante["nome_completo"],
                "cpf": participante["cpf"],
                "email": participante["email"],
                "telefone": participante["numero"],
                "data_nascimento": participante["data_nascimento"],
                "problemas": problemas,
            })

    return render_template(
        "auditoria_inconsistencias.html",
        inconsistencias=inconsistencias,
        total_participantes=len(participantes),
    )


# =========================
# ARRECADAÇÕES
# =========================
@app.route("/participantes/<int:id>/arrecadar", methods=["POST"])
def salvar_arrecadacao(id):
    if "user_id" not in session:
        return redirect("/login")

    try:
        valor = parse_currency_input(request.form["valor"])
    except ValueError:
        flash("Informe um valor de arrecadação válido.", "danger")
        return redirect(f"/participantes/{id}")

    data_lancamento = request.form["data_lancamento"]
    observacao = request.form["observacao"]

    try:
        validate_iso_date(data_lancamento, "Data de lançamento")
    except ValueError as error:
        flash(str(error), "danger")
        return redirect(f"/participantes/{id}")

    conn = get_db()
    conn.execute("""
        INSERT INTO arrecadacoes (participante_id, valor, data_lancamento, observacao)
        VALUES (?, ?, ?, ?)
    """, (id, valor, data_lancamento, observacao))
    conn.commit()
    conn.close()

    flash("Arrecadação registrada com sucesso!", "success")
    return redirect(f"/participantes/{id}")


@app.route("/arrecadacoes/<int:id>/editar", methods=["GET", "POST"])
def editar_arrecadacao(id):
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()
    arrecadacao = conn.execute("""
        SELECT * FROM arrecadacoes
        WHERE id = ?
    """, (id,)).fetchone()

    if not arrecadacao:
        conn.close()
        return "Registro não encontrado"

    if request.method == "POST":
        try:
            valor = parse_currency_input(request.form["valor"])
        except ValueError:
            conn.close()
            flash("Informe um valor de arrecadação válido.", "danger")
            return redirect(f"/arrecadacoes/{id}/editar")

        data_lancamento = request.form["data_lancamento"]
        observacao = request.form["observacao"]

        try:
            validate_iso_date(data_lancamento, "Data de lançamento")
        except ValueError as error:
            conn.close()
            flash(str(error), "danger")
            return redirect(f"/arrecadacoes/{id}/editar")

        conn.execute("""
            UPDATE arrecadacoes
            SET valor = ?, data_lancamento = ?, observacao = ?
            WHERE id = ?
        """, (valor, data_lancamento, observacao, id))

        conn.commit()
        participante_id = arrecadacao["participante_id"]
        conn.close()

        flash("Arrecadação atualizada com sucesso!", "success")
        return redirect(f"/participantes/{participante_id}")

    conn.close()
    return render_template("editar_arrecadacao.html", arrecadacao=arrecadacao)


@app.route("/arrecadacoes/<int:id>/excluir", methods=["POST"])
def excluir_arrecadacao(id):
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()
    registro = conn.execute("""
        SELECT participante_id FROM arrecadacoes WHERE id = ?
    """, (id,)).fetchone()

    if not registro:
        conn.close()
        return "Registro não encontrado"

    participante_id = registro["participante_id"]
    conn.execute("DELETE FROM arrecadacoes WHERE id = ?", (id,))
    conn.commit()
    conn.close()

    flash("Arrecadação excluída com sucesso!", "success")
    return redirect(f"/participantes/{participante_id}")


# =========================
# RELATÓRIOS
# =========================
@app.route("/relatorios")
def relatorios():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    congregacao = request.args.get("congregacao", "").strip()
    macro = request.args.get("macro", "").strip()

    filtro = "WHERE 1=1"
    params = []

    if congregacao:
        filtro += " AND participantes.congregacao = ?"
        params.append(congregacao)

    if macro:
        filtro += """
            AND participantes.congregacao IN (
                SELECT nome FROM congregacoes
                WHERE macro_id = (SELECT id FROM macros WHERE nome = ?)
            )
        """
        params.append(macro)

    participantes_relatorio = conn.execute(f"""
        SELECT
            participantes.id,
            participantes.nome_completo,
            participantes.data_nascimento,
            participantes.cpf,
            participantes.email,
            participantes.numero,
            participantes.nome_mae,
            participantes.congregacao,
            macros.nome AS macro,
            COALESCE(SUM(arrecadacoes.valor), 0) AS total_arrecadado
        FROM participantes
        LEFT JOIN congregacoes ON congregacoes.nome = participantes.congregacao
        LEFT JOIN macros ON macros.id = congregacoes.macro_id
        LEFT JOIN arrecadacoes ON arrecadacoes.participante_id = participantes.id
        {filtro}
        GROUP BY
            participantes.id,
            participantes.nome_completo,
            participantes.data_nascimento,
            participantes.cpf,
            participantes.email,
            participantes.numero,
            participantes.nome_mae,
            participantes.congregacao,
            macros.nome
        ORDER BY participantes.nome_completo ASC
    """, params).fetchall()

    por_macro = conn.execute(f"""
        SELECT macros.nome AS macro, COALESCE(SUM(arrecadacoes.valor), 0) AS total
        FROM macros
        LEFT JOIN congregacoes ON congregacoes.macro_id = macros.id
        LEFT JOIN participantes ON participantes.congregacao = congregacoes.nome
        LEFT JOIN arrecadacoes ON arrecadacoes.participante_id = participantes.id
        WHERE 1=1
          AND (? = '' OR macros.nome = ?)
          AND (? = '' OR participantes.congregacao = ?)
        GROUP BY macros.id, macros.nome
        ORDER BY macros.nome
    """, (macro, macro, congregacao, congregacao)).fetchall()

    macros = conn.execute("SELECT * FROM macros ORDER BY nome").fetchall()
    congregacoes = conn.execute("""
        SELECT congregacoes.nome, macros.nome AS macro_nome
        FROM congregacoes
        JOIN macros ON macros.id = congregacoes.macro_id
        ORDER BY macros.nome, congregacoes.nome
    """).fetchall()

    total_geral = conn.execute(f"""
        SELECT COALESCE(SUM(arrecadacoes.valor), 0) AS total
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        {filtro}
    """, params).fetchone()["total"]

    conn.close()

    return render_template(
        "relatorios.html",
        participantes_relatorio=participantes_relatorio,
        por_macro=por_macro,
        macros=macros,
        congregacoes=congregacoes,
        total_geral=total_geral
    )


@app.route("/relatorios/planilha")
def relatorios_planilha():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    congregacao = request.args.get("congregacao", "").strip()
    macro = request.args.get("macro", "").strip()

    filtro = "WHERE 1=1"
    params = []

    if congregacao:
        filtro += " AND participantes.congregacao = ?"
        params.append(congregacao)

    if macro:
        filtro += """
            AND participantes.congregacao IN (
                SELECT nome FROM congregacoes
                WHERE macro_id = (SELECT id FROM macros WHERE nome = ?)
            )
        """
        params.append(macro)

    participantes_relatorio = conn.execute(f"""
        SELECT
            participantes.id,
            participantes.nome_completo,
            COALESCE(macros.nome, '-') AS macro,
            participantes.congregacao,
            participantes.data_nascimento,
            participantes.cpf,
            participantes.email,
            participantes.numero,
            participantes.nome_mae,
            COALESCE(SUM(arrecadacoes.valor), 0) AS total_arrecadado
        FROM participantes
        LEFT JOIN congregacoes ON congregacoes.nome = participantes.congregacao
        LEFT JOIN macros ON macros.id = congregacoes.macro_id
        LEFT JOIN arrecadacoes ON arrecadacoes.participante_id = participantes.id
        {filtro}
        GROUP BY
            participantes.id,
            participantes.nome_completo,
            macros.nome,
            participantes.congregacao,
            participantes.data_nascimento,
            participantes.cpf,
            participantes.email,
            participantes.numero,
            participantes.nome_mae
        ORDER BY participantes.nome_completo ASC
    """, params).fetchall()

    total_geral = conn.execute(f"""
        SELECT COALESCE(SUM(arrecadacoes.valor), 0) AS total
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        {filtro}
    """, params).fetchone()["total"]

    por_macro = conn.execute(f"""
        SELECT COALESCE(macros.nome, 'Sem macro') AS macro, COALESCE(SUM(arrecadacoes.valor), 0) AS total
        FROM participantes
        LEFT JOIN congregacoes ON congregacoes.nome = participantes.congregacao
        LEFT JOIN macros ON macros.id = congregacoes.macro_id
        LEFT JOIN arrecadacoes ON arrecadacoes.participante_id = participantes.id
        {filtro}
        GROUP BY macros.nome
        ORDER BY macros.nome
    """, params).fetchall()

    conn.close()

    workbook = Workbook()
    ws_resumo = workbook.active
    ws_resumo.title = "Resumo"
    ws_dados = workbook.create_sheet("Participantes")
    ws_resumo.sheet_view.showGridLines = False
    ws_dados.sheet_view.showGridLines = False
    ws_resumo.sheet_view.zoomScale = 115
    ws_dados.sheet_view.zoomScale = 90
    ws_resumo.sheet_properties.tabColor = "1D4ED8"
    ws_dados.sheet_properties.tabColor = "0F172A"
    ws_resumo.freeze_panes = "A9"

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    total_participantes = len(participantes_relatorio)
    total_macros = len([item for item in por_macro if float(item["total"] or 0) > 0])
    media_arrecadacao = float(total_geral or 0) / total_participantes if total_participantes else 0

    titulo_font = Font(name="Calibri", size=17, bold=True, color="FFFFFF")
    subtitulo_font = Font(name="Calibri", size=10, italic=True, color="DBEAFE")
    cabecalho_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    texto_font = Font(name="Calibri", size=11, color="1E293B")
    texto_destaque_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    card_label_font = Font(name="Calibri", size=10, bold=True, color="475569")
    card_value_font = Font(name="Calibri", size=15, bold=True, color="0F172A")
    moeda_font = Font(name="Calibri", size=11, bold=True, color="166534")
    moeda_card_font = Font(name="Calibri", size=16, bold=True, color="166534")
    linklike_font = Font(name="Calibri", size=11, bold=True, color="1D4ED8")
    fill_titulo = PatternFill(fill_type="solid", start_color="1D4ED8", end_color="1D4ED8")
    fill_cabecalho = PatternFill(fill_type="solid", start_color="0F172A", end_color="0F172A")
    fill_info = PatternFill(fill_type="solid", start_color="EFF6FF", end_color="EFF6FF")
    fill_info_label = PatternFill(fill_type="solid", start_color="DBEAFE", end_color="DBEAFE")
    fill_resumo_total = PatternFill(fill_type="solid", start_color="DCFCE7", end_color="DCFCE7")
    fill_linha_clara = PatternFill(fill_type="solid", start_color="F8FAFC", end_color="F8FAFC")
    fill_linha_branca = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    fill_total_coluna = PatternFill(fill_type="solid", start_color="F0FDF4", end_color="F0FDF4")
    fill_card = PatternFill(fill_type="solid", start_color="F8FAFC", end_color="F8FAFC")
    fill_card_value = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    border_destaque = Border(
        left=Side(style="medium", color="93C5FD"),
        right=Side(style="medium", color="93C5FD"),
        top=Side(style="medium", color="93C5FD"),
        bottom=Side(style="medium", color="93C5FD"),
    )

    def estilizar_intervalo(worksheet, start_row, end_row, start_col, end_col, fill=None, font=None, alignment=None, cell_border=None):
        for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                if fill is not None:
                    cell.fill = fill
                if font is not None:
                    cell.font = font
                if alignment is not None:
                    cell.alignment = alignment
                if cell_border is not None:
                    cell.border = cell_border

    # ABA RESUMO
    ws_resumo.merge_cells("A1:H1")
    ws_resumo["A1"] = "Relatorio de Arrecadacao"
    ws_resumo["A1"].font = titulo_font
    ws_resumo["A1"].fill = fill_titulo
    ws_resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_resumo.merge_cells("A2:H2")
    ws_resumo["A2"] = "Visao executiva com indicadores e consolidado por macro"
    ws_resumo["A2"].font = subtitulo_font
    ws_resumo["A2"].fill = fill_titulo
    ws_resumo["A2"].alignment = Alignment(horizontal="center", vertical="center")

    cards = [
        ("A4:B4", "A5:B6", "Total arrecadado", float(total_geral or 0), moeda_card_font, "R$ #,##0.00", fill_resumo_total),
        ("C4:D4", "C5:D6", "Participantes", total_participantes, card_value_font, None, fill_card_value),
        ("E4:F4", "E5:F6", "Macros com arrecadacao", total_macros, card_value_font, None, fill_card_value),
        ("G4:H4", "G5:H6", "Media por participante", media_arrecadacao, moeda_card_font, "R$ #,##0.00", fill_card_value),
    ]

    for label_range, value_range, label, value, value_font, number_format, value_fill in cards:
        ws_resumo.merge_cells(label_range)
        ws_resumo.merge_cells(value_range)
        label_cell = ws_resumo[label_range.split(":")[0]]
        value_cell = ws_resumo[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = value
        label_cell.font = card_label_font
        value_cell.font = value_font
        label_cell.fill = fill_card
        value_cell.fill = value_fill
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if number_format:
            value_cell.number_format = number_format
        estilizar_intervalo(ws_resumo, label_cell.row, value_cell.row + 1, label_cell.column, value_cell.column + 1, cell_border=border_destaque)

    ws_resumo["A8"] = "Macro"
    ws_resumo["B8"] = "Total"
    ws_resumo["A8"].font = cabecalho_font
    ws_resumo["B8"].font = cabecalho_font
    ws_resumo["A8"].fill = fill_cabecalho
    ws_resumo["B8"].fill = fill_cabecalho
    ws_resumo["A8"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo["B8"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo["A8"].border = border
    ws_resumo["B8"].border = border

    linha_resumo = 9
    for item in por_macro:
        ws_resumo[f"A{linha_resumo}"] = item["macro"]
        ws_resumo[f"B{linha_resumo}"] = float(item["total"] or 0)
        ws_resumo[f"B{linha_resumo}"].number_format = "R$ #,##0.00"
        linha_resumo += 1

    ultima_linha_macro = max(9, linha_resumo - 1)
    for row in ws_resumo.iter_rows(min_row=9, max_row=ultima_linha_macro, min_col=1, max_col=2):
        for cell in row:
            cell.font = moeda_font if cell.column == 2 else texto_font
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if cell.column == 2 else "left", vertical="center")
            cell.fill = fill_linha_clara if cell.row % 2 == 0 else fill_linha_branca

    ws_resumo["E8"] = "Filtros aplicados"
    ws_resumo["E8"].font = cabecalho_font
    ws_resumo["E8"].fill = fill_cabecalho
    ws_resumo["E8"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.merge_cells("E9:H9")
    ws_resumo["E9"] = f"Macro: {macro or 'Todas'} | Congregacao: {congregacao or 'Todas'} | Gerado em: {agora}"
    ws_resumo["E9"].font = texto_font
    ws_resumo["E9"].fill = fill_info
    ws_resumo["E9"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    estilizar_intervalo(ws_resumo, 8, 9, 5, 8, cell_border=border)

    if por_macro:
        grafico_macro = BarChart()
        grafico_macro.type = "bar"
        grafico_macro.style = 10
        grafico_macro.title = "Arrecadacao por Macro"
        grafico_macro.y_axis.title = "Valor"
        grafico_macro.x_axis.title = "Macro"
        grafico_macro.height = 8
        grafico_macro.width = 14
        dados_grafico = Reference(ws_resumo, min_col=2, min_row=8, max_row=ultima_linha_macro)
        categorias_grafico = Reference(ws_resumo, min_col=1, min_row=9, max_row=ultima_linha_macro)
        grafico_macro.add_data(dados_grafico, titles_from_data=True)
        grafico_macro.set_categories(categorias_grafico)
        grafico_macro.legend = None
        ws_resumo.add_chart(grafico_macro, "D11")

    for coluna, largura in {
        "A": 20,
        "B": 18,
        "C": 16,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
    }.items():
        ws_resumo.column_dimensions[coluna].width = largura
    ws_resumo.row_dimensions[1].height = 30
    ws_resumo.row_dimensions[2].height = 22
    for linha in range(4, 7):
        ws_resumo.row_dimensions[linha].height = 24

    # ABA DE DADOS (PLANILHA PRINCIPAL)
    ws_dados.merge_cells("A1:J1")
    ws_dados["A1"] = "Planilha de Participantes"
    ws_dados["A1"].font = titulo_font
    ws_dados["A1"].fill = fill_titulo
    ws_dados["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_dados.merge_cells("A2:J2")
    ws_dados["A2"] = f"Macro: {macro or 'Todas'}  |  Congregacao: {congregacao or 'Todas'}  |  Gerado em: {agora}"
    ws_dados["A2"].font = subtitulo_font
    ws_dados["A2"].fill = fill_titulo
    ws_dados["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws_dados.merge_cells("A3:J3")
    ws_dados["A3"] = f"Resumo rapido: {total_participantes} participantes listados e total arrecadado de R$ {float(total_geral or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    ws_dados["A3"].font = texto_destaque_font
    ws_dados["A3"].fill = fill_info
    ws_dados["A3"].alignment = Alignment(horizontal="center", vertical="center")
    estilizar_intervalo(ws_dados, 1, 3, 1, 10, cell_border=border)

    ws_dados.row_dimensions[1].height = 30
    ws_dados.row_dimensions[2].height = 22
    ws_dados.row_dimensions[3].height = 24

    headers = [
        "ID",
        "Nome",
        "Macro",
        "Congregacao",
        "Data nascimento",
        "CPF",
        "E-mail",
        "Numero",
        "Nome da mae",
        "Total arrecadado",
    ]

    header_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws_dados.cell(row=header_row, column=col, value=header)
        cell.font = cabecalho_font
        cell.fill = fill_cabecalho
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    data_start = header_row + 1
    for idx, p in enumerate(participantes_relatorio, start=data_start):
        ws_dados.cell(row=idx, column=1, value=p["id"])
        ws_dados.cell(row=idx, column=2, value=p["nome_completo"])
        ws_dados.cell(row=idx, column=3, value=p["macro"])
        ws_dados.cell(row=idx, column=4, value=p["congregacao"])
        ws_dados.cell(row=idx, column=5, value=p["data_nascimento"])
        ws_dados.cell(row=idx, column=6, value=p["cpf"])
        ws_dados.cell(row=idx, column=7, value=p["email"] or "-")
        ws_dados.cell(row=idx, column=8, value=p["numero"] or "-")
        ws_dados.cell(row=idx, column=9, value=p["nome_mae"] or "-")
        total_cell = ws_dados.cell(row=idx, column=10, value=float(p["total_arrecadado"] or 0))
        total_cell.number_format = "R$ #,##0.00"

    possui_dados = len(participantes_relatorio) > 0
    last_data_row = data_start + len(participantes_relatorio) - 1 if possui_dados else header_row

    if possui_dados:
        for row in ws_dados.iter_rows(min_row=data_start, max_row=last_data_row, min_col=1, max_col=10):
            for cell in row:
                cell.font = texto_font
                cell.border = border
                cell.fill = fill_linha_clara if cell.row % 2 == 0 else fill_linha_branca
                if cell.column == 10:
                    cell.font = moeda_font
                    cell.fill = fill_total_coluna
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif cell.column == 2:
                    cell.font = texto_destaque_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif cell.column == 3:
                    cell.font = linklike_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column in (1, 5):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        total_row = last_data_row + 2
        ws_dados[f"I{total_row}"] = "Total geral"
        ws_dados[f"I{total_row}"].font = texto_destaque_font
        ws_dados[f"I{total_row}"].fill = fill_info_label
        ws_dados[f"I{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dados[f"I{total_row}"].border = border_destaque
        ws_dados[f"J{total_row}"] = f"=SUM(J{data_start}:J{last_data_row})"
        ws_dados[f"J{total_row}"].font = moeda_card_font
        ws_dados[f"J{total_row}"].fill = fill_resumo_total
        ws_dados[f"J{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws_dados[f"J{total_row}"].border = border_destaque
        ws_dados[f"J{total_row}"].number_format = "R$ #,##0.00"
    else:
        ws_dados.merge_cells("A6:J6")
        ws_dados["A6"] = "Nenhum participante encontrado para os filtros selecionados."
        ws_dados["A6"].font = texto_destaque_font
        ws_dados["A6"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dados["A6"].border = border
        ws_dados["A6"].fill = fill_info
        last_data_row = 6

    ws_dados.freeze_panes = "A6"

    tabela_ref = f"A5:J{last_data_row}"
    if possui_dados:
        tabela = Table(displayName="TabelaParticipantes", ref=tabela_ref)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws_dados.add_table(tabela)
    else:
        ws_dados.auto_filter.ref = "A5:J5"

    larguras = {
        "A": 8,
        "B": 36,
        "C": 14,
        "D": 26,
        "E": 16,
        "F": 18,
        "G": 32,
        "H": 16,
        "I": 30,
        "J": 18,
    }
    for col, largura in larguras.items():
        ws_dados.column_dimensions[col].width = largura

    xlsx_output = BytesIO()
    workbook.save(xlsx_output)
    xlsx_output.seek(0)

    response = make_response(xlsx_output.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=relatorio_participantes.xlsx"
    return response

@app.route("/relatorios/pdf")
def relatorios_pdf():
    if "user_id" not in session:
        return redirect("/login")

    conn = get_db()

    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()
    congregacao = request.args.get("congregacao", "").strip()
    macro = request.args.get("macro", "").strip()

    filtro = "WHERE 1=1"
    params = []

    if data_inicio:
        filtro += " AND data_lancamento >= ?"
        params.append(data_inicio)

    if data_fim:
        filtro += " AND data_lancamento <= ?"
        params.append(data_fim)

    if congregacao:
        filtro += " AND participantes.congregacao = ?"
        params.append(congregacao)

    if macro:
        filtro += """
            AND participantes.congregacao IN (
                SELECT nome FROM congregacoes
                WHERE macro_id = (SELECT id FROM macros WHERE nome = ?)
            )
        """
        params.append(macro)

    total = conn.execute(f"""
        SELECT COALESCE(SUM(valor), 0) AS total
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        {filtro}
    """, params).fetchone()["total"]

    total_lancamentos = conn.execute(f"""
        SELECT COUNT(*) AS total
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        {filtro}
    """, params).fetchone()["total"]

    por_congregacao = conn.execute(f"""
        SELECT participantes.congregacao, COALESCE(SUM(valor), 0) AS total
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        {filtro}
        GROUP BY participantes.congregacao
        ORDER BY total DESC
    """, params).fetchall()

    por_macro = conn.execute(f"""
        SELECT macros.nome AS macro, COALESCE(SUM(valor), 0) AS total
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        JOIN congregacoes ON congregacoes.nome = participantes.congregacao
        JOIN macros ON macros.id = congregacoes.macro_id
        {filtro}
        GROUP BY macros.nome
        ORDER BY total DESC
    """, params).fetchall()

    detalhes = conn.execute(f"""
        SELECT participantes.nome_completo, participantes.congregacao, arrecadacoes.valor, arrecadacoes.data_lancamento, arrecadacoes.observacao
        FROM arrecadacoes
        JOIN participantes ON participantes.id = arrecadacoes.participante_id
        {filtro}
        ORDER BY arrecadacoes.data_lancamento DESC, arrecadacoes.id DESC
    """, params).fetchall()

    conn.close()

    html = f"""
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{
            font-family: Arial, sans-serif;
            font-size: 12px;
            color: #111;
        }}

        .topo {{
            text-align: center;
            margin-bottom: 20px;
            border-bottom: 2px solid #ccc;
            padding-bottom: 10px;
        }}

        h1, h2, h3 {{
            margin: 0 0 10px 0;
        }}

        .resumo {{
            margin-bottom: 20px;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }}

        th, td {{
            border: 1px solid #ccc;
            padding: 6px;
            text-align: left;
        }}

        th {{
            background: #f2f2f2;
        }}

        .small {{
            color: #555;
            font-size: 11px;
        }}
    </style>
</head>

<body>

    <div class="topo">
        <h1>Sistema Cadepa</h1>
        <h3>Relatório de Arrecadação</h3>
        <p class="small">Sistema de gestão do congresso</p>
    </div>

    <div class="resumo">
        <h3>Resumo</h3>
        <p><strong>Total arrecadado:</strong> R$ {total:.2f}</p>
        <p><strong>Total de lançamentos:</strong> {total_lancamentos}</p>
        <p><strong>Filtro data início:</strong> {data_inicio or '-'}</p>
        <p><strong>Filtro data fim:</strong> {data_fim or '-'}</p>
        <p><strong>Filtro macro:</strong> {macro or '-'}</p>
        <p><strong>Filtro congregação:</strong> {congregacao or '-'}</p>
    </div>

    <h3>Detalhamento das Arrecadações</h3>

    <table>
        <tr>
            <th>Participante</th>
            <th>Congregação</th>
            <th>Valor</th>
            <th>Data</th>
            <th>Observação</th>
        </tr>
"""

    for d in detalhes:
        html += f"""
        <tr>
            <td>{d['nome_completo']}</td>
            <td>{d['congregacao']}</td>
            <td>R$ {d['valor']:.2f}</td>
            <td>{d['data_lancamento']}</td>
            <td>{d['observacao'] or '-'}</td>
        </tr>
"""

    html += """
        </table>
    </body>
    </html>
    """

    pdf = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf)

    if pisa_status.err:
        return "Erro ao gerar PDF"

    response = make_response(pdf.getvalue())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "attachment; filename=relatorio_arrecadacao.pdf"
    return response

@app.route("/demo", methods=["GET", "POST"])
def demo():
    if request.method != "POST":
        return redirect("/login")

    if not DEMO_PASSWORD:
        flash("Modo demonstração está desativado neste ambiente.", "warning")
        return redirect("/login")

    senha_demo = request.form.get("senha_demo", "")
    if senha_demo != DEMO_PASSWORD:
        flash("Senha do modo demonstração inválida.", "danger")
        return redirect("/login")

    session["user_id"] = 999
    session["nome"] = "Modo Demonstração"
    session["cargo"] = "demo"
    session["modo_demo"] = True
    session.permanent = True
    return redirect("/dashboard")


# =========================
# RUN
# =========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=not IS_PRODUCTION)